from fastapi import FastAPI, Response
from pydantic import BaseModel
from typing import List, Optional, Union
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment

app = FastAPI()

class Item(BaseModel):
    sku: Optional[str] = ""
    desc: Optional[str] = ""
    qty: Optional[Union[int, float, str, None]] = None
    price: Optional[Union[int, float, str, None]] = None

class InvoicePayload(BaseModel):
    invoice_number: str = "Invoice"
    issue_date: str = ""
    quote_invoice: str = "Quote"
    purchase_number: Optional[str] = ""
    customer: dict = {}
    client_contact: Optional[str] = ""
    project_name: Optional[str] = ""
    meta: dict = {}
    items: List[Item] = []

class HtmlPayload(BaseModel):
    html: str

# ---- helpers ----
def build_xlsx_from_json(data: InvoicePayload) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    cust = data.customer or {}
    meta = data.meta or {}
    # header block
    ws["A1"] = data.quote_invoice or "Quote"
    ws["A3"] = f"Project: {data.project_name or ''}"
    ws["A4"] = f"Customer: {cust.get('name','')}"
    ws["A5"] = f"Address: {cust.get('address','')}"
    ws["A6"] = f"Contact: {data.client_contact or ''}"
    ws["A7"] = f"Email: {cust.get('email','')}"
    ws["E3"] = f"Quote No: {data.invoice_number or ''}"
    ws["E4"] = f"Date: {data.issue_date or ''}"
    ws["E5"] = f"PO: {meta.get('po','')}"
    ws["E6"] = f"Currency: {meta.get('currency','')}"

    # columns
    widths = [6, 50, 12, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[{1:'A',2:'B',3:'C',4:'D',5:'E'}[i]].width = w

    start = 10
    headers = ["#", "Description", "Qty", "Unit Price", "Amount"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h); cell.border = border

    r = start + 1
    for idx, it in enumerate(data.items or [], start=1):
        desc = " — ".join([s for s in [(it.sku or "").strip(), (it.desc or "").strip()] if s])
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=desc).alignment = wrap
        ws.cell(row=r, column=3, value=it.qty)
        ws.cell(row=r, column=4, value=it.price)
        # if qty & price numeric, compute amount; else echo price (for totals rows)
        try:
            amt = (float(it.qty) if it.qty not in (None,"") else None)
            price = (float(it.price) if isinstance(it.price,(int,float,str)) and str(it.price).replace(',','').replace('£','').strip().replace('.','',1).isdigit() else None)
            ws.cell(row=r, column=5, value=(amt*price if (amt is not None and price is not None) else it.price))
        except:
            ws.cell(row=r, column=5, value=it.price)
        for c in range(1,6):
            ws.cell(row=r, column=c).border = border
        r += 1

    # pad blank rows so next section starts at row 90
    FIRST_DYNAMIC_ROW = 90
    while r < FIRST_DYNAMIC_ROW:
        for c in range(1,6):
            ws.cell(row=r, column=c).border = border
        r += 1

    # simple footer placeholders
    ws.cell(row=r,   column=4, value="VAT @ 20%")
    ws.cell(row=r+1, column=4, value="Total")

    bio = BytesIO(); wb.save(bio); return bio.getvalue()

def build_xlsx_from_html(html_str: str) -> bytes:
    # pick first table in HTML; requires pandas only (bundled in most server images; add to requirements)
    dfs = pd.read_html(html_str)
    df = dfs[0] if dfs else pd.DataFrame()
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    return bio.getvalue()

# ---- endpoints ----
@app.post("/invoice.xlsx")
def invoice_xlsx(payload: InvoicePayload):
    xls = build_xlsx_from_json(payload)
    fname = f"{payload.invoice_number or 'Invoice'}.xlsx"
    return Response(content=xls,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@app.post("/html-to-excel.xlsx")
def html_to_excel_xlsx(payload: HtmlPayload):
    xls = build_xlsx_from_html(payload.html)
    return Response(content=xls,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="table.xlsx"'})
